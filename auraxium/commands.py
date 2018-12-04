import asyncio

import discord
from openpyxl import Workbook, load_workbook

# All functions defined this module will be accessible to bot users.

# Initialize Excel worksheets
try:
    wb = load_workbook('data/notifications.xlsx')
except FileNotFoundError:
    wb = Workbook()
    wb.save('data/notifications.xlsx')
sheet_notifications = wb.active


# async def census(client, msg, words, help=False):
# """Generic command for looking up information about ingame data."""

# Warn users that more specific commands exist
# (weaponinfo, outfitroster, playerinfo, etc.)

# ?census player/character/char Auroram

# Player name
# Battle rank & ASP
# outfit
# Empire
# Server
# 3 top weapons
# Primary roles (add role if playtime(30d) over 10h?) or just top 2-3

# async def player(client, msg, words, help=False):
#     """Displays basic information about a player."""
#
#     _battle_rank = "17"
#     _asp = " (ASP)"
#     _faction_icon_url = ""
#     _kdr = "2.05"
#
#     _outfit_name =
#     _outfit_tag =
#     _outfit_join_date =
#     _outfit_rank =
#     _outfit_rank_ordinal =
#
#     _top_weapon_1_name =
#     _top_weapon_1_type =
#     _top_weapon_1_kills =
#     _top_weapon_2_name =
#     _top_weapon_2_type =
#     _top_weapon_2_kills =
#     _top_weapon_3_name =
#     _top_weapon_3_type =
#     _top_weapon_3_kills =
#
#     _top_domain_1_name =
#     _top_domain_1_hours =
#     _top_domain_2_name =
#     _top_domain_2_hours =
#     _top_domain_3_name =
#     _top_domain_3_hours =
#
#     embed = discord.Embed(title="Auroram", colour=discord.Colour(
#         0x34f4d6), description="Vanu Sovereignty combatant on Cobalt.")
#
#     embed.set_thumbnail(url=_faction_icon_url)
#     embed.set_author(name="Player information:")
#     embed.set_footer(text="\"Look, Mega - no hands!\"")
#
#     embed.add_field(name="Basic information", value="Battle rank: 17 (ASP)\nKDR: 2.04".format(
#         _battle_rank, _asp, _kdr))
#     embed.add_field(
#         name="Outfit", value="Members of the Utopian Mummy (MUMS)\nJoined 08/10/2018.\nRank: \"Son\" (tier 5).")
#     embed.add_field(name="Top weapons",
#                     value="#1 - C4 - [Explosive] - 2406 kills\n#2 - Force blade - [Knife] - 1206 kills\n#3 - Pulsar LSW - [LMG] - 1202 kills")
#     embed.add_field(name="Most played",
#                     value="Combat Medic - 300 hours\nHeavy Assault - 260 hours\nSunderer - 120 hours")
#
#     await bot.say(content="Here is the information you requested, {}:".format(msg.author.mention), embed=embed)


async def echo(client, msg, words, help=False):
    """Repeats the rest of the user's message."""

    await client.send_message(msg.channel, content=msg.content[5:])


async def shutdown(client, msg, words, help=False):
    """Ends the asyncio loop."""

    await client.send_message(msg.channel, content='Shutting down... bye!')
    await client.close()


async def dinnerTime(client, msg, words, help=False):
    """Moves members into the voice channel of the user sending the command.

    Any user that has been mentioned in this message and is connected to a
    voice channel in the current server will be moved to the voice channel of
    the user who sent the command.
    """

    # Check permissions
    role = discord.utils.get(msg.server.roles, name='Son')
    if msg.author.top_role < role:
        content = ('Sorry {}, you need to have the rank of {} or higher to '
                   'use this command.').format(msg.author.mention, role.name)
        await client.send_message(msg.channel, content=content)
        return

    # Get the channel to move members into
    if msg.author.voice.voice_channel == None:
        content = '{} You are not connected to a voice channel.'.format(
            msg.author.mention)
        await client.send_message(msg.channel, content=content)
        return
    else:
        author_channel = msg.author.voice.voice_channel

    # Create a list of all users that have been mentioned in this message
    mentioned_users = [m for m in msg.server.members if m.mentioned_in(msg)]

    # Get a list of all non-afk members in voice channels
    voice_users = [m for c in msg.server.channels if len(
        c.voice_members) > 0 for m in c.voice_members]

    # Get a list of all members that are both mentioned and in voice channels
    members_to_move = [
        m for m in mentioned_users if m in voice_users and not m in author_channel.voice_members and not m.voice.is_afk]

    if len(members_to_move) == 0:
        content = '{} I was unable to find members that can be moved to {}.'.format(
            msg.author.mention, author_channel.name)
        await client.send_message(msg.channel, content=content)
        return
    elif len(members_to_move) == 1:
        content = '{} Moving 1 member to {}.'.format(
            msg.author.mention, author_channel.name)
        await client.send_message(msg.channel, content=content)
    else:
        content = '{} Moving {} members to {}.'.format(
            msg.author.mention, len(members_to_move), author_channel.name)
        await client.send_message(msg.channel, content=content)

    for member in members_to_move:
        await client.move_member(member, author_channel)


async def juicy(client, msg, words, help=False):
    """Forwards any message send to the user Spidey to the #living-room channel.
    These two values are hard-coded as I do not see any reason to extend this
    functionality.
    """

    # Hard-coding, best coding!
    DEST_CHANNEL = client.get_channel('483582158993096704')
    USER_SPIDEY = await client.get_user_info('243477509520359427')

    # Help text
    if help:
        return ('Any message prefixed with this command will be '
                'forwarded to the {} channel while also mentioning '
                'Spidey.').format(DEST_CHANNEL.mention)

    # If the author does not have an avatar set, use their respective default
    author_icon_url = msg.author.avatar_url
    if author_icon_url == '':
        author_icon_url = msg.author.default_avatar_url

    # Generate the embed
    embed = discord.Embed(colour=discord.Colour(0x8452ae),
                          description=msg.content[6:],
                          timestamp=msg.timestamp)
    embed.set_author(name=msg.author.display_name,
                     icon_url=author_icon_url)
    embed.set_footer(text='# {}'.format(msg.channel.name))

    # Send the message
    await client.send_message(DEST_CHANNEL,
                              content=USER_SPIDEY.mention,
                              embed=embed)


# async def notify(client, msg, words, help=False):
#     """Allows users to notify their absence for upcoming Ops.
#     Valid parameters are "absent", "late" and "clear".
#     """
#
#     # Convert the author to the format "Name#0123"
#     discord_user = str(msg.author)
#
#     # Fake names
#     sheet_notifications['A4'].value = 'Leon#0201'
#     sheet_notifications['B4'].value = 'Notified absent'
#
#     # Does the Discord user already exist in the worksheet?
#     user_list = [cell.value for cell in sheet_notifications['A']]
#     row = user_list.index(discord_user) + 1
#     if row != 0:
#         user_row = [cell.value for cell in sheet_notifications[row:row]]
#         print(user_row)

    # if words[0] == 'absent':
    #     # Mark the Discord user as notified absent
    #     if user_row >= 0:
    #         if user_row
    #     if discord_user in name_col:
    #         if user_row == 'Notified absent':
    #             # Tell them that you already marked them
    #             reply = 'You are already notified absent.'
    #         else:
    #             # Mark them and thank them for telling you
    #             reply = 'Copy that, I marked you as notified absent.'
    #
    # elif words[0] == 'late':
    #     # Mark the user as notified late
    #     if value == 'Notified late':
    #         # Tell them that you already marked them
    #         reply = 'You are already notified late.'
    #     else:
    #         # Mark them and thank them for telling you
    #         reply = 'Copy that, I marked you as notified late.'
    #
    # elif words[0] == 'clear':
    #     # Clear any notifications for this user
    #     if name_exists:
    #         # Tell them you removed their entry
    #         reply = 'Copy that, I cleared any notifications for you.'
    #     else:
    #         # Tell them there is nothing to remove
    #         reply = 'There are no notifications for you.'
    #
    # else:
    #     # Complain about usage
    #     reply = ('Wait, what?\nPlease make sure you only write `?notify '
    #              'absent|late|clear` and nothing else.')


async def ping(client, msg, words, help=False):
    await client.send_message(msg.channel, content='{} Pong!'.format(msg.author.mention))


async def OPStimeLegacy(client, msg, words, help=False):
    """Moves members into the voice channel of the user sending the command.

    Any user that has been mentioned in this message and is connected to a
    voice channel in the current server will be moved to the voice channel of
    the user who sent the command.
    """




async def steal_that_bastards_avatar(client, msg, words, help=False):
    """Retrieves a user based on their discord name and posts their avatar."""

    help_text = 'Usage: ?steal_that_bastards_avatar CHANNEL_MENTION USER_MENTION'

    # Get the user
    user = words[0]

    # The number of messages to scan
    messages = 1000

    # username#discriminator
    name, discriminator = words[0].split('#')

    # Cycle through the messages
    user_hash = ''
    async for message in client.logs_from(msg.channel, limit=messages):
        if message.author.name == name and message.author.discriminator == discriminator:
            user_hash = message.author.id
            break

    # If no user was found
    if user_hash == '':
        content = ('{} I was unable to find a user matching the '
                   'criteria.').format(msg.author.mention)

        await client.send_message(msg.channel, content=content)
        return
    else:
        user = await client.get_user_info(user_hash)
        content = '{} I found a match: {}.'.format(msg.author.mention,
                                                   user.avatar_url)
        await client.send_message(msg.channel, content=content)
