import logging
from os.path import isfile

from openpyxl import Workbook, load_workbook

# Create a logger
logger = logging.getLogger('auraxium.sheets')

# Default path, prefixed to all document paths
base_path = 'data/'

# The global list of documents.
documents = []


# This entire module is an abstraction of openpyxl functionality.
# I found the original syntax awkward to use and have added my
# own classes and helper functions to make using it easier for my
# application.
# Especially functions like "append_row(list)" were sorely missing
# from the original code.


class Cell(object):
    """Represents a single cell in a worksheet.
    """

    def __init__(self, sheet, row=None, column=None, string=None):
        self.column = column
        self._op = None
        self.row = row
        self.sheet = sheet
        self._value = None

        # If both row and column have been set
        if self.column != None and self.row != None:
            self._op = sheet._op.cell(column=self.column, row=self.row)

        # If not, see if the string has been set
        else:
            # try:
            self._op = sheet._op[str]
            self.column = self._op.col_idx
            self.row = self._op.row
            # except WhateverError:

        logger.debug('Created Cell with row: {} and column: {}.'.format(
            self.row, self.column))

    @property
    def value(self):
        """The value of the cell."""
        self._value = self._op.value
        return self._value

    @value.setter
    def value(self, new_value):
        self._op.value = new_value
        self._value = new_value
        self.sheet._update()

    @value.deleter
    def value(self):
        # TODO: Unset the value in openpyxl
        self._value = None


class Document(object):
    """Abstraced version of an Excel Workbook.
    """

    def __init__(self, name, autosave=True):
        self.autosave = autosave
        self._filepath = '{}{}.xlsx'.format(base_path, name)
        self.name = name
        self._op = None
        self.sheets = []

        # Try loading the file
        try:
            logger.debug('Creating new document "{}".'.format(name))
            self._op = load_workbook(self._filepath)

        # The file is in use or otherwise unavailable. The changes made will
        # be saved to a different file.
        except PermissionError:
            old_path = self._filepath

            i = 1
            while True:
                new_path = '{}{}_{}.xlsx'.format(base_path, name, i)
                if not isfile(new_path):
                    self._filepath = new_path
                    break
                i += 1

            self._op = Workbook()
            self._op.save(self._filepath)
            logger.warning(('Cannot access file "{}". Saving to "{}" '
                            'instead.').format(old_path, new_path))

        # The file does not exist and needs to be created.
        except FileNotFoundError:
            self._op = Workbook()
            self._op.save(self._filepath)
            logger.debug(
                'File "{}" not found but created.'.format(self._filepath))

        # Add the document to the global document list
        documents.append(self)


class Range(object):
    """A range.
    height is calculated as the entire height, the "col/row" reference is
    the top left corner of the range
    """

    def __init__(self, sheet, row=None, column=None, width=1, height=1):
        self.column = column
        self.height = height
        self.row = row
        self.sheet = sheet
        self._values = []
        self.width = width

    @property
    def values(self):
        """A 2D-list of values for this range."""
        self._values = [[cell for cell in row]
                        for row in self.sheet._op.values]
        # self._values = [cell for row in self.sheet._op.values for cell in row]
        return self._values

    @values.setter
    def values(self, new_values):
        if isinstance(new_values[0], list):
            new_values_2 = new_values
        else:
            new_values_2 = [new_values]
        # new_values is expected to be an array of at most the size of the
        # range itself.
        row_x = 0
        for r_ in self.sheet._op.iter_rows(min_row=self.row,
                                           max_row=self.row + self.height - 1,
                                           min_col=self.column,
                                           max_col=self.column + self.width - 1):
            row_x += 1
            for cell in r_:
                cell.value = new_values_2[row_x - 1][cell.col_idx - 1]
        self._values = new_values_2
        self.sheet._update()

    @values.deleter
    def values(self):
        for r_ in self.sheet._op.iter_rows(min_row=self.row,
                                           max_row=self.row + self.height - 1,
                                           min_col=self.column,
                                           max_col=self.column + self.width - 1):
            for cell in r_:
                cell.value = None
        self._values = [[]]


class Sheet(object):
    def __init__(self, name, document=None):
        self.document = document
        self.name = name
        self._op = None

        # If no document has been specified, add it to the default document,
        # "data/Auraxium.xlsx"
        if self.document == None:
            self.document = Document('auraxium')
            self._op = self.document._op.active
            self._op.title = name
        else:
            self._op = self.document._op.create_sheet(name)
        self.save()

    def append_row(self, values):
        """Fills the first final (?) empty row with the values from the list.
        """
        self.range(self._op.max_row + 1, 1, len(values)).values = values

    def cell(self, row=None, column=None, string=None):
        return Cell(self, row, column, string)

    def delete_row(self, row):
        self._op.delete_rows(row, amount=1)
        self._update()

    def find_in_column(self, column, value):
        """Searches a column for a given value and returns the row
        the value first appears in.
        """
        values = [cell[column - 1] for cell in self._op.values]
        try:
            return values.index(value) + 1
        except ValueError:
            return 0

    def range(self, row, column, width=1, height=1):
        return Range(self, row, column, width, height)

    def save(self, save_as=None):
        logger.debug('Saving file "{}..."'.format(self.document.name))
        self.document._op.save(self.document._filepath)

    def _update(self):
        if self.document.autosave:
            self.save()


def get_document(name):
    """Returns a document object by its name."""
    document = [d for d in documents if d.name == name]
    if len(document) > 0:
        return document[0]
